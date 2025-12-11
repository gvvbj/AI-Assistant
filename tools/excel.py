import openpyxl
from tools.registry import tool_registry
from utils.security import SecurityManager
from utils.error_handling import safe_execute

@tool_registry.register(
    name="excel_delete",
    description="Delete rows from Excel.",
    parameters={
        "type": "object",
        "properties": {
            "file_path": {"type": "string"},
            "row_index": {"type": "string", "description": "Rows to delete, e.g. '2,3'"},
            "target_content": {"type": "string", "description": "Ignored, kept for compatibility"}
        },
        "required": ["file_path"]
    }
)
@safe_execute("Excel删除失败")
def delete_excel_rows(file_path, row_index=None, target_content=None):
    # 1. 路径安全检查
    clean_path = SecurityManager.sanitize_path(file_path)
    
    # 2. 业务逻辑
    wb = openpyxl.load_workbook(clean_path)
    sheet = wb.active
    
    if row_index:
        # 倒序删除防止索引偏移
        idxs = sorted([int(i) for i in str(row_index).split(",") if i.strip().isdigit()], reverse=True)
        count = 0
        for i in idxs:
            sheet.delete_rows(i)
            count += 1
        wb.save(clean_path)
        return f"已删除 {count} 行"
    
    return "未提供 row_index"

@tool_registry.register(
    name="excel_read",
    description="Read Excel data.",
    parameters={
        "type": "object",
        "properties": {
            "file_path": {"type": "string"},
            "query": {"type": "string"}
        },
        "required": ["file_path"]
    }
)
@safe_execute("Excel读取失败")
def read_excel(file_path, query=None):
    clean_path = SecurityManager.sanitize_path(file_path)
    wb = openpyxl.load_workbook(clean_path, data_only=True)
    sheet = wb.active
    data = []
    # 简单读取前20行
    for row in sheet.iter_rows(max_row=20, values_only=True):
        data.append(str(row))
    return "\n".join(data)


@tool_registry.register(
    name="excel_write",  # 这里的名字必须和模型调用的名字完全一致
    description="Write or append data to an Excel file.",
    parameters={
        "type": "object",
        "properties": {
            "file_path": {
                "type": "string", 
                "description": "Path to the excel file"
            },
            "data": {
                "type": "array", 
                "items": {"type": "string"},
                "description": "A list of data to append as a new row, e.g. ['Tom', '18', 'Student']"
            },
            "sheet_name": {
                "type": "string",
                "description": "Optional sheet name, default is active sheet"
            }
        },
        "required": ["file_path", "data"]
    }
)
@safe_execute("Excel写入失败")
def write_excel_row(file_path, data, sheet_name=None):
    # 1. 路径安全清洗
    clean_path = SecurityManager.sanitize_path(file_path)
    
    # 2. 加载工作簿 (注意：这里不能用 read_only=True，因为要写入)
    wb = openpyxl.load_workbook(clean_path)
    
    # 3. 选择工作表
    if sheet_name and sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
    else:
        sheet = wb.active
        
    # 4. 追加数据 (openpyxl 的 append 方法可以直接接受列表)
    # 注意：如果 data 是字符串形式的列表 "[a,b]"，可能需要先转成 list
    if isinstance(data, str):
        # 简单的字符串转列表处理，防止模型传错类型
        import ast
        try:
            data = ast.literal_eval(data)
        except:
            # 如果转换失败，就把字符串当做单列数据
            data = [data]
            
    sheet.append(data)
    
    # 5. 保存文件
    wb.save(clean_path)
    
    return f"成功写入数据: {data}"